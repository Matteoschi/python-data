import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
import random

# File path and sheet name
FOLDER = "tournament_manager"
os.makedirs(FOLDER, exist_ok=True)
FILE = 'tournament.xlsx'
FILE_PATH = os.path.join(FOLDER, FILE)
SHEET_NAME = 'Players'

# List to store player names
player_list = []

def add_players():
    try:
        num_players = int(input("How many players are there? "))
    except ValueError:
        print("❌ Please enter a valid number.")
        return add_players()
    
    for i in range(num_players):
        while True:
            player = input(f"Enter name of player {i+1}: ").lower().strip()
            if player:
                if player in player_list:
                    print("⚠️ Player already registered. Choose a different name.")
                    continue
                print("✅ Player added.")
                player_list.append(player)
                break
            else:
                print("❌ Invalid name. Try again.")

    print(f"Registered players: {player_list}")

def read_players(registered_players):
    print("👥 Registered players:")
    for i, name in enumerate(registered_players, 1):
        print(f"{i}: {name}")

def register_match(registered_players):
    if len(registered_players) < 2:
        print("❌ At least two registered players are required to record a match.")
        return

    match_players = []
    print("✍️ Enter names of 2 players. Press Enter to stop.")

    for i in range(2):
        player_name = input("Player name: ").lower().strip()
        if not player_name:
            if len(match_players) < 2:
                print("❌ You must enter two players.")
                continue
            else:
                break
        if player_name not in [p.lower() for p in registered_players]:
            print("❌ Player not found.")
            continue
        if player_name in match_players:
            print("⚠️ Player already entered.")
            continue
        match_players.append(player_name)

    comment = ""
    if input("Do you want to add a match description? (y/n): ").strip().lower() == "y":
        comment = input("Enter comment: ").strip()

    score = input("Enter the result (e.g. 3-2): ").strip()
    try:
        score_a, score_b = map(int, score.split("-"))
        print(f"✅ Result: {score_a}-{score_b}")
    except ValueError:
        print("❌ Invalid score format. Use 'number-number' (e.g. 3-2).")

    date = datetime.now().strftime("%Y-%m-%d %H:%M")
    versus = " vs ".join(match_players)

    wb = load_workbook(FILE_PATH)

    if "Matches" not in wb.sheetnames:
        ws_matches = wb.create_sheet("Matches")
        ws_matches.append(["Date", "Player 1", "Player 2", "Score 1", "Score 2", "Score", "Comment"])
    else:
        ws_matches = wb["Matches"]

    ws_matches.append([date, match_players[0], match_players[1], score_a, score_b, score, comment])
    wb.save(FILE_PATH)
    print(f"✅ Match recorded: {versus} ({score})")
    if comment:
        print(f"📝 Comment saved: {comment}")

def add_new_players(registered_players):
    new_players = []
    try:
        n = int(input("🔢 How many new players do you want to add? "))
    except ValueError:
        print("❌ Please enter a valid number.")
        return

    for i in range(n):
        while True:
            name = input(f"Enter name for player {i+1}: ").strip().lower()
            if not name:
                print("❌ Invalid name. Try again.")
                continue
            if name in registered_players or name in new_players:
                print("⚠️ Player already exists. Choose another.")
                continue
            new_players.append(name)
            break

    if not new_players:
        print("❌ No players added.")
        return

    for name in new_players:
        ws.append([name])
    
    if len(new_players) > 1:
        print(f"✅ {len(new_players)} players added successfully.")
    else:
        print(f"✅ 1 player added successfully.")

    wb.save(FILE_PATH)

def update_statistics():
    wb = load_workbook(FILE_PATH)

    if "Statistics" in wb.sheetnames:
        del wb["Statistics"]
    ws_stats = wb.create_sheet("Statistics")
    print("Updating statistics sheet...")

    matches = []
    if "Matches" not in wb.sheetnames:
        print("❌ No matches recorded.")
        return

    ws_matches = wb["Matches"]
    for row in ws_matches.iter_rows(min_row=2, values_only=True):
        try:
            _, p1, p2, s1, s2, _, _ = row
            p1, p2 = p1.lower(), p2.lower()
            s1, s2 = int(s1), int(s2)
            matches.append((p1, p2, s1, s2))
        except (ValueError, TypeError):
            continue

    if not matches:
        print("❌ No valid matches found.")
        return

    stats = {}
    for p1, p2, s1, s2 in matches:
        for p in [p1, p2]:
            if p not in stats:
                stats[p] = {
                    "played": 0, "won": 0, "lost": 0, "drawn": 0,
                    "points_for": 0, "points_against": 0
                }

        stats[p1]["played"] += 1
        stats[p2]["played"] += 1

        stats[p1]["points_for"] += s1
        stats[p1]["points_against"] += s2
        stats[p2]["points_for"] += s2
        stats[p2]["points_against"] += s1

        if s1 > s2:
            stats[p1]["won"] += 1
            stats[p2]["lost"] += 1
        elif s2 > s1:
            stats[p2]["won"] += 1
            stats[p1]["lost"] += 1
        else:
            stats[p1]["drawn"] += 1
            stats[p2]["drawn"] += 1

    ws_stats.append([
        "Player", "Games Played", "Won", "Lost", "Drawn", "Points For", "Points Against", "Win Rate (%)"
    ])

    for player, s in stats.items():
        win_rate = round((s["won"] / s["played"]) * 100, 1) if s["played"] > 0 else 0.0
        drawn = s["drawn"] if s["drawn"] > 0 else "NA"
        ws_stats.append([
            player.capitalize(), s["played"], s["won"], s["lost"],
            drawn, s["points_for"], s["points_against"], win_rate
        ])
        print(f"\n👤 {player.capitalize()}")
        print(f"  • Games played: {s['played']}")
        print(f"  • Won: {s['won']}, Lost: {s['lost']}, Drawn: {drawn}")
        print(f"  • Points for: {s['points_for']}, Against: {s['points_against']}")
        print(f"  • Win rate: {win_rate}%")

    wb.save(FILE_PATH)
    print("\n📁 Statistics sheet updated and saved.")

def torneo_eliminazione(squadre):
    wb = load_workbook(FILE_PATH)

    # Gestione foglio "torneo"
    if "torneo" in wb.sheetnames:
        eliminare = input("Vuoi creare un nuovo torneo? (s/n): ").lower().strip()
        if eliminare == "s":
            del wb["torneo"]
            print("✅ Vecchio torneo eliminato.")
    if "torneo" not in wb.sheetnames:
        ws_torneo = wb.create_sheet("torneo")
        ws_torneo.append(["Data", "Giocatore 1", "Giocatore 2", "Punto 1", "Punto 2", "Risultato", "Vincitore", "Commento"])
    else:
        ws_torneo = wb["torneo"]

    turno = 1
    while len(squadre) > 1:
        print(f"\n🎯 --- Turno {turno} - {len(squadre)} giocatori ---")
        random.shuffle(squadre)
        vincitori = []

        squadra_bye = None
        if len(squadre) % 2 != 0:
            squadra_bye = squadre.pop()  # Rimuove e salva l'ultimo
            vincitori.append(squadra_bye)
            print(f"\n⚠️ {squadra_bye} PASSA AUTOMATICAMENTE IL TURNO")
            date = datetime.now().strftime("%Y-%m-%d %H:%M")
            ws_torneo.append([date, squadra_bye, "-", "-", "-", "-", squadra_bye, "Passa automaticamente il turno"])

        for i in range(0, len(squadre), 2):
            squadra1 = squadre[i]
            squadra2 = squadre[i + 1]
            print(f"\n🏁 {squadra1} vs {squadra2}")

            # Inserimento vincitore
            while True:
                vincitore = input("👉 Inserisci il vincitore: ").strip().lower()
                if vincitore != squadra1.lower() and vincitore != squadra2.lower():
                    print("❌ Errore: il vincitore deve essere uno dei due.")
                else:
                    vincitore = vincitore.capitalize()
                    break

            # Inserimento risultato
            while True:
                risultato = input("📊 Inserisci il risultato (es. 2-1): ").strip()
                try:
                    g1, g2 = map(int, risultato.split("-"))
                except ValueError:
                    print("❌ Formato non valido. Usa es. 2-1")
                    continue

                if g1 == g2:
                    print("❌ Pareggi non ammessi nel torneo a eliminazione.")
                else:
                    break

            # Commento opzionale
            vuole_commento = input("💬 Inserisci un commento (oppure digita 'n'): ").strip()
            commento = "" if vuole_commento.lower() == "n" else vuole_commento

            date = datetime.now().strftime("%Y-%m-%d %H:%M")
            ws_torneo.append([
                date, squadra1, squadra2, g1, g2, risultato, vincitore, commento
            ])

            print(f"✅ Vince {vincitore} con il risultato di {g1}-{g2}")
            vincitori.append(vincitore)

        squadre = vincitori
        turno += 1
    vittoriosa = squadre[0]
    print(f"\n🏆 La squadra vincitrice del torneo è: {vittoriosa.upper()}")
    ws_torneo.append([date, "-", "-", "-", "-", vittoriosa.upper(), "VINCITORE"])
    wb.save(FILE_PATH)

# === Avvio del programma ===
if not os.path.exists(FILE_PATH):
    print("🛠️ Nessun file trovato. Creazione nuovo file Excel...")
    
    add_players()  # Richiede input per aggiungere giocatori

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Inserimento intestazione e giocatori nella prima colonna
    ws.append(["Players"])
    for player in player_list:
        ws.append([player])

    wb.save(FILE_PATH)
    print(f"✅ File '{FILE}' creato nella cartella '{FOLDER}' con i giocatori registrati.")

    registered_players = player_list.copy()

else:
    print(f"📂 Il file '{FILE}' esiste già nella cartella '{FOLDER}'.")
    print("📥 Apertura file in corso...")

    try:
        wb = load_workbook(FILE_PATH)
        ws = wb[SHEET_NAME]
        print("✅ File aperto con successo.")

        registered_players = [
            row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]
        ]

        if not registered_players:
            print("⚠️ Nessun giocatore trovato nel file.")
    except Exception as e:
        print(f"❌ Errore durante l'apertura del file: {e}")
        registered_players = []

# === Ciclo principale ===
while True:
    num_players = len(registered_players)

    if num_players % 2 == 0 and num_players > 0:
        print(f"\n✅ È possibile giocare il torneo: ci sono {num_players} giocatori.")
        print("Opzioni disponibili:")
        print("[1=Aggiungi giocatori, 2=Lista giocatori, 3=Registra partita, 4=Statistiche, 5=Torneo a eliminazione, 0=Esci]")
    else:
        print(f"\n⚠️ Numero di giocatori non valido per il torneo (attualmente {num_players}).")
        print("Opzioni disponibili:")
        print("[1=Aggiungi giocatori, 2=Lista giocatori, 3=Registra partita, 4=Statistiche, 0=Esci]")

    action = input("Cosa vuoi fare? ").strip()

    if action == "1":
        add_new_players(registered_players)

        # 🔄 Aggiorna la lista dei giocatori dopo aggiunta
        wb = load_workbook(FILE_PATH)
        ws = wb[SHEET_NAME]
        registered_players = [
            row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]
        ]

    elif action == "2":
        read_players(registered_players)

    elif action == "3":
        register_match(registered_players)
        wb = load_workbook(FILE_PATH)  # Ricarica il file per aggiornamenti

    elif action == "4":
        update_statistics()

    elif action == "5":
        if num_players % 2 == 0 and num_players > 0:
            torneo_eliminazione(registered_players)
        else:
            print("❌ Il torneo può iniziare solo con un numero pari di giocatori maggiore di zero.")

    elif action == "0":
        print("👋 Uscita dal programma.")
        break

    else:
        print("❌ Opzione non valida. Riprova.")