import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from collections import defaultdict
folder = r'spliter'
os.makedirs(folder, exist_ok=True)
file = 'dati.xlsx'
file_name = os.path.join(folder, file)
sheet_name = 'Registro'
intestazioni = ['Data e Ora', 'Persona', 'Cifra', "Descrizione"]

if not os.path.exists(file_name):
    # FILE NON ESISTE: chiedi i nomi dei partecipanti
    partecipanti = int(input("Quanti partecipanti siete? "))
    nomi = []
    for i in range(partecipanti):
        while True:
            nome = input(f"Inserisci il nome del partecipante {i + 1}: ").strip()
            if nome:
                nomi.append(nome)
                break
            else:
                print("⚠️ Nome non valido. Riprova.") 

    print("I partecipanti sono:", nomi)

    wb = Workbook() # Crea un nuovo workbook (file Excel vuoto).
    ws = wb.active #Ottiene il foglio attivo
    ws.title = sheet_name # Imposta il nome del foglio (tab)
    for col, val in enumerate(intestazioni, start=1): # Scrive le intestazioni nella prima riga (riga 1, colonne A, B, C, …).
        ws.cell(row=1, column=col, value=val)

    # Salva subito per creare il file e intestazioni
    wb.save(file_name)
    print(f"File {file} creato nella cartella: {folder}")

    for nome in nomi:
        # Scrivi riga fittizia nel foglio Excel con cifra = 0
        ora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.append([ora, nome, 0.0, "Inizializzazione partecipante"])

else:
    # FILE ESISTE: leggi i nomi già presenti nel file Excel
    wb = load_workbook(file_name) # Apre un file Excel già esistente 
    ws = wb[sheet_name]
    if [cell.value for cell in ws[1]] != intestazioni: # Controlla se la prima riga del foglio (riga 1) contiene le intestazioni corrette.ws[1] è la prima riga, e [cell.value for cell in ws[1]] crea una lista con i valori delle celle.Se sono diversi da quelli attesi (intestazioni), allora li aggiorna.
        for col, val in enumerate(intestazioni, start=1):
            ws.cell(row=1, column=col, value=val)
        wb.save(file_name)
        print(f"{file}: intestazioni aggiornate.")
    else:
        print(f"Aperto {file} regolarmente")

    # Estrai la lista nomi unici dal file (colonna Persona)
    nomi_set = set() # Crea un set vuoto, cioè un contenitore che non permette duplicati.
    for row in ws.iter_rows(min_row=2, values_only=True):# Scorre tutte le righe del foglio ws, partendo dalla seconda riga (min_row=2), perché la prima contiene le intestazioni.
        nome = row[1]
        if nome:
            nomi_set.add(nome)
    nomi = sorted(list(nomi_set)) # Converte il set in una lista e la ordina in ordine alfabetico.
    print("Partecipanti trovati nel file:", nomi)

# LEGGI SPESE PRECEDENTI DAL FILE EXCEL (per calcolo totale)
totali_precedenti = defaultdict(float)
if ws.max_row > 1:
    for row in ws.iter_rows(min_row=2, values_only=True): # Scorre tutte le righe del foglio ws, partendo dalla seconda riga (min_row=2), perché la prima contiene le intestazioni.
        nome = row[1]
        cifra = row[2]
        if nome in nomi and cifra is not None:
            totali_precedenti[nome] += cifra

# INSERIMENTO NUOVE SPESE
spese = []
while True:
    spesa_nome = input("\nChi ha pagato? (scrivi 'fine' per terminare) ").strip()
    if spesa_nome.lower() == "fine":
        break
    if spesa_nome not in nomi:
        aggiungi = input(f"⚠️ Il nome '{spesa_nome}' non è presente. Vuoi aggiungerlo? (s/n) ").strip().lower()
        if aggiungi == 's':
            nomi.append(spesa_nome)
            print(f"✅ Aggiunto {spesa_nome} alla lista dei partecipanti.")
            
            # Scrivi riga fittizia nel foglio Excel con cifra = 0
            ora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws.append([ora, spesa_nome, 0.0, "Inizializzazione partecipante"])
            wb.save(file_name)
        else:
            continue
    while True:
        try:
            quanto = float(input("Quanto ha pagato? (in euro) "))
            break
        except ValueError:
            print("⚠️ Inserisci un numero valido.")
    per_cosa = input("Per cosa ha pagato? ").strip()

    spese.append({
        "nome": spesa_nome,
        "quanto": quanto,
        "descrizione": per_cosa
    })

    # Se file esiste e più nomi, continua a chiedere spese fino a 'fine'
    # Se solo 1 nome, chiedi altre spese o termina?
    if len(nomi) == 1:
        continua = input("Vuoi inserire un'altra spesa? (s/n) ").strip().lower()
        if continua != 's':
            break

# CALCOLI TOTALE + QUOTA
totali = {nome: totali_precedenti[nome] for nome in nomi}
for spesa in spese:
    totali[spesa["nome"]] += spesa["quanto"]

totale = sum(totali.values())
quota = totale / len(nomi) if nomi else 0

print(f"\nTotale speso: {totale:.2f} €")
print(f"Quota per partecipante: {quota:.2f} €\n")

for nome in nomi:
    print(f"{nome} ha speso {totali[nome]:.2f} €")

saldo = {nome: totali[nome] - quota for nome in nomi}

print("\n💸 Saldo finale:")
for nome in nomi:
    print(f"{nome}: {saldo[nome]:+.2f} €")

# Chi deve a chi
creditori = [(nome, s) for nome, s in saldo.items() if s > 0]
debitori = [(nome, -s) for nome, s in saldo.items() if s < 0]

# CREA/AGGIORNA IL FOGLIO RIEPILOGO
riepilogo_name = 'Riepilogo'
if riepilogo_name in wb.sheetnames:
    ws_r = wb[riepilogo_name]
    wb.remove(ws_r)  # Rimuovi vecchio riepilogo per riscrivere da capo
ws_r = wb.create_sheet(riepilogo_name)

# Scrivi intestazioni
ws_r.append(["Info", "Valore"])

# Totale speso e quota
ws_r.append(["Totale speso", totale])
ws_r.append(["Quota per partecipante", quota])

# Spesa per partecipante
ws_r.append([])
ws_r.append(["Spesa per partecipante"])
ws_r.append(["Nome", "Spesa"])
for nome in nomi:
    ws_r.append([nome, totali[nome]])

# Saldo finale
ws_r.append([])
ws_r.append(["Saldo finale"])
ws_r.append(["Nome", "Saldo"])
for nome in nomi:
    ws_r.append([nome, saldo[nome]])

# Chi deve a chi
ws_r.append([])
ws_r.append(["Chi deve a chi"])
ws_r.append(["Debitore", "Deve dare", "Creditore"])

i, j = 0, 0
while i < len(debitori) and j < len(creditori):
    deb_nome, deb_amm = debitori[i]
    cred_nome, cred_amm = creditori[j]
    pagamento = min(deb_amm, cred_amm)
    ws_r.append([deb_nome, pagamento, cred_nome])
    deb_amm -= pagamento
    cred_amm -= pagamento
    debitori[i] = (deb_nome, deb_amm)
    creditori[j] = (cred_nome, cred_amm)
    if deb_amm == 0:
        i += 1
    if cred_amm == 0:
        j += 1

# Salva il file (già presente nel tuo codice)
wb.save(file_name)


# SALVA NUOVE SPESE
for voce in spese:
    ora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.append([ora, voce["nome"], voce["quanto"], voce["descrizione"]])

# Salva il file una sola volta dopo tutte le modifiche
wb.save(file_name)
print(f"\nDati salvati in {file_name}")